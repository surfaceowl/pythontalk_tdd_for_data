# file encoding: utf-8
"""
functions analyzing document and workbook content excel worksheet
"""
import logging
import os
import re

import openpyxl
import openpyxl.styles.colors

from src.configuration_info_cais import list_section_names


def config_logging():
    logging.basicConfig(filename="./log_cais.txt",
                        filemode='a',
                        format='%(message)s',
                        level=logging.DEBUG)
    logging.getLogger().setLevel(logging.INFO)
    logger = logging.getLogger('')
    logger.handlers = []


def find_xlsx_files(filepath):
    """"
    creates list of all .xlsx files in data path
    :param filepath: path to data directory
    :type filepath: string
    :return files: list of filenames that are *.xlsx files
    """
    files = [file for file in os.listdir(filepath) if
             (not file.startswith('~') and file.endswith(".xlsx"))]  # ignore open/temp files
    return files


def find_school_year_in_filename(name):
    """
    finds year and school year embedded in name of file if it exists in filename
    school_year taken as the range of YYYY-YYYY, year taken as the first YYYY segment
    :param name: filename of input file
    :type name: string
    :return: year, school_year
    :rtype: basestring
    """
    match = re.findall(r"(\d\d\d\d)", name, re.IGNORECASE)
    if match:
        year = int(match[0])
        school_year = str(match[0]) + str("-") + str(match[1])
    else:
        year, school_year = "Not Found", "Not Found"

    year = year or None
    school_year = school_year or None

    return year, school_year


def get_workbook(excel_file):
    """
    function to open excel file input and return openpyxl object
    :param excel_file: input
    :return: workbook; openpyxl object
    :rtype: object
    """
    try:
        workbook = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        raise FileNotFoundError
    except FileExistsError:
        raise FileExistsError
    except OSError:
        raise OSError
    return workbook


def get_sheet_names(workbook):
    """
    gets list of sheet names and identifies locations of key info
    :param workbook: input excel file
    :return: list of lists;
    list_sheets - list of sheet names (what you see in excel)
    sheet_index_members - list of sheets with member schools only: sheet #, sheet name, value in
    1st row
    sheet_index_heads - list of sheets with Head of School names,
    sheet_index_directors - list of sheets with Director names
    """

    # identify sections by making array of sheets and content titles
    # to associate content sections with specific sheets, so we can search for certain content types
    # in subsets of the workbook (e.g. members are on tabs 2-24 for workbook `x`)
    list_sheets = workbook.sheetnames
    sheet_index_all = []

    for i, worksheet in enumerate(workbook.sheetnames):
        sheet_index_all.append(
            [i, str(workbook[worksheet].title), workbook[worksheet].cell(row=1, column=1).value])

    member_start, head_start, director_start = None, None, None
    sheet_index_members = None
    sheet_index_heads, sheet_index_directors = None, None
    len_all_sheets = len(sheet_index_all)

    for i, item in enumerate(sheet_index_all):
        regex_member = re.escape(list_section_names[0])
        regex_head = re.escape(list_section_names[2])
        regex_director = re.escape(list_section_names[3])
        find_member = re.search(regex_member, str(item), re.IGNORECASE)
        find_head = re.search(regex_head, str(item), re.IGNORECASE)
        find_director = re.search(regex_director, str(item), re.IGNORECASE)

        if find_member:
            member_start = i
        if find_head:
            head_start = i
        if find_director:
            director_start = i

    # after checking all sheets, if key values are still None - they are not included
    # most important job is to get complete range for members (sheet_index_members)
    if member_start is not None and head_start is not None:
        sheet_index_members = sheet_index_all[member_start:head_start]
    elif member_start is not None and head_start is None:
        # assume heads and directors are missing, and members are on all tabs
        logging.warning("`DIRECTORS` section not found. Guessing END range for sheet_index_members")
        sheet_index_members = sheet_index_all[member_start:len_all_sheets]
    elif member_start is None and head_start is not None:
        # case - author did not include section title 'MEMBERS"
        logging.warning("`MEMBERS` section not found. Guessing range for sheet_index_members")
        sheet_index_members = sheet_index_all[0:head_start]
    elif member_start is None and head_start is None:
        # case - author forgot all section titles
        # include all sheets & see what data extraction can find
        logging.warning("`MEMBERS` nor `HEAD` sections found.")
        sheet_index_members = sheet_index_all[0:len_all_sheets]

    if head_start is not None and director_start is not None:
        sheet_index_heads = sheet_index_all[head_start:director_start]
    elif head_start is not None and director_start is None:
        sheet_index_heads = sheet_index_all[head_start:len_all_sheets]
    elif head_start is None and director_start is not None:
        logging.warning("`HEAD` section not found, but `DIRECTOR` section found.")
        pass
    elif head_start is None and director_start is None:
        logging.warning("Neither `HEAD` nor `DIRECTORS` sections found.")
        pass

    if director_start is not None:
        sheet_index_directors = sheet_index_all[director_start]
    else:
        logging.warning("`DIRECTORS` section not found.")
        pass

    # logging.info(f"list sheets = {list_sheets}")
    # logging.info(f"sheet_index_members = {sheet_index_members}")
    # logging.info(f"sheet_index_heads = {sheet_index_heads}")
    # logging.info(f"sheet_index_directors = {sheet_index_directors}")

    return list_sheets, sheet_index_members, sheet_index_heads, sheet_index_directors


if __name__ == '__main__':
    config_logging()

    path_testing = "./data/test_cais/"
    target_file = path + "School_Directory_2013-2014-converted.xlsx"
    target_file2 = path_testing + "School_Directory_2018-2019-converted.xlsx"
